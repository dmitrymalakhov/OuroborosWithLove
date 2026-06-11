"""Excel template tools.

These tools complement document analysis: documents.py reads spreadsheets as
documents, while this module inspects and fills workbook templates.
"""

from __future__ import annotations

import datetime
import json
import pathlib
import re
from typing import Any, Dict, Iterable, List, Tuple

from ouroboros.tools.registry import ToolContext, ToolEntry
from ouroboros.utils import safe_relpath

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_XLSX_BYTES = 50 * 1024 * 1024
MAX_SCAN_ROWS = 300
MAX_SCAN_COLS = 80
MAX_INSPECT_FORMULAS = 200
MAX_INPUT_CANDIDATES = 250
MIN_WRITE_CONFIDENCE = 0.75
MAX_CHART_SERIES = 12
MAX_CHART_POINTS = 1000

_CELL_RE = re.compile(r"^\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6}$", re.IGNORECASE)
_FORMULA_REF_RE = re.compile(
    r"(?:(?:'((?:[^']|'')+)'|([A-Za-z_][A-Za-z0-9_ .]*))!)?"
    r"(\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6})"
    r"(?::(\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6}))?",
    re.IGNORECASE,
)
_RANGE_RE = re.compile(
    r"^(?:(?:'((?:[^']|'')+)'|([^!]+))!)?(\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6}"
    r"(?::\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6})?)$",
    re.IGNORECASE,
)


def _require_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import range_boundaries

        return load_workbook, range_boundaries
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required for Excel template tools. Install dependencies from requirements.txt."
        ) from exc


def _scope(ctx: ToolContext) -> Dict[str, Any]:
    return {
        "user_id": ctx.current_user_id,
        "user_role": ctx.user_role,
        "drive_root": str(ctx.drive_root),
        "shared_drive_root": str(ctx.shared_drive_root or ctx.drive_root),
    }


def _resolve_workspace_file(ctx: ToolContext, path: str, source: str = "drive") -> Tuple[pathlib.Path, pathlib.Path]:
    if not path or not isinstance(path, str):
        raise ValueError("path must be a non-empty string")

    source = str(source or "drive").strip().lower()
    if source not in {"drive", "repo"}:
        raise ValueError("source must be 'drive' or 'repo'")
    if source == "repo" and str(ctx.user_role or "user").lower() != "admin":
        raise PermissionError("source='repo' is admin-only in multi-user mode")

    root = (ctx.repo_dir if source == "repo" else ctx.drive_root).resolve()
    resolved = (root / safe_relpath(path)).resolve()
    try:
        resolved.relative_to(root)
    except ValueError:
        raise ValueError("Path traversal is not allowed")
    if not resolved.exists():
        raise FileNotFoundError(f"File not found: {path}")
    if not resolved.is_file():
        raise ValueError(f"Path is not a file: {path}")
    return resolved, root


def _resolve_xlsx_path(ctx: ToolContext, path: str, source: str = "drive") -> Tuple[pathlib.Path, pathlib.Path]:
    resolved, root = _resolve_workspace_file(ctx, path, source)
    if resolved.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported in the first spreadsheet tool version.")
    size = resolved.stat().st_size
    if size > MAX_XLSX_BYTES:
        raise ValueError(f"Workbook is too large: {size} bytes")
    return resolved, root


def _safe_output_path(
    root: pathlib.Path,
    output_path: str,
    template_name: str,
    overwrite: bool,
    suffix_label: str = "filled",
) -> Tuple[pathlib.Path, str]:
    if output_path:
        rel = safe_relpath(output_path)
        if pathlib.PurePosixPath(rel).suffix.lower() != ".xlsx":
            raise ValueError("output_path must end with .xlsx")
    else:
        stamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%d-%H%M%S")
        stem = _safe_stem(pathlib.Path(template_name).stem)
        suffix_label = _safe_stem(suffix_label) or "filled"
        rel = str(pathlib.PurePosixPath("spreadsheets") / f"{stem}-{suffix_label}-{stamp}.xlsx")

    path = (root / rel).resolve()
    try:
        path.relative_to(root.resolve())
    except ValueError:
        raise ValueError("Path traversal is not allowed")
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists() and not overwrite:
        stem = path.stem
        suffix = path.suffix
        parent_rel = pathlib.PurePosixPath(rel).parent
        counter = 2
        while path.exists():
            rel = str(parent_rel / f"{stem}-{counter}{suffix}")
            path = (root / rel).resolve()
            counter += 1
    return path, rel


def _safe_sheet_title(value: str, fallback: str = "Chart") -> str:
    value = re.sub(r"[\[\]:*?/\\]+", "-", str(value or fallback)).strip()
    value = value.strip("'")[:31]
    return value or fallback


def _unique_sheet_title(wb: Any, desired: str) -> str:
    base = _safe_sheet_title(desired)
    if base not in wb.sheetnames:
        return base
    for idx in range(2, 100):
        suffix = f" {idx}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in wb.sheetnames:
            return candidate
    raise ValueError("Could not create a unique chart sheet name")


def _safe_stem(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9А-Яа-я._ -]+", "-", str(value or "template"))
    value = re.sub(r"\s+", "-", value.strip(" .-_"))
    return value[:80] or "template"


def _cell_is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _cell_is_writable(ws: Any, cell: Any) -> bool:
    if ws.protection.sheet and getattr(getattr(cell, "protection", None), "locked", False):
        return False
    return not _cell_is_formula(cell.value)


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, datetime.date):
        return value.isoformat()
    return str(value)


def _normalize_coordinate(value: str) -> str:
    return str(value or "").replace("$", "").upper()


def _split_range_ref(default_sheet: str, ref: str) -> Tuple[str, str]:
    raw = str(ref or "").strip()
    if raw.startswith("="):
        raw = raw[1:]
    raw = raw.replace("$", "")
    match = _RANGE_RE.match(raw)
    if not match:
        raise ValueError(f"Invalid range reference: {ref!r}")
    sheet = match.group(1) or match.group(2) or default_sheet
    sheet = str(sheet or "").strip().strip("'").replace("''", "'")
    target = _normalize_coordinate(match.group(3))
    if not sheet:
        raise ValueError(f"Range reference must include a sheet or data_sheet: {ref!r}")
    return sheet, target


def _range_dimensions(target: str, range_boundaries: Any) -> Tuple[int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(target)
    rows = max_row - min_row + 1
    cols = max_col - min_col + 1
    cells = rows * cols
    return rows, cols, cells


def _quote_sheet_for_formula(sheet: str) -> str:
    return "'" + str(sheet).replace("'", "''") + "'"


def _strip_formula_string_literals(formula: str) -> str:
    text = str(formula or "")
    result = []
    in_string = False
    idx = 0
    while idx < len(text):
        char = text[idx]
        if char == '"':
            in_string = not in_string
            result.append(" ")
            idx += 1
            continue
        result.append(" " if in_string else char)
        idx += 1
    return "".join(result)


def _formula_dependencies(formula: str, current_sheet: str) -> List[str]:
    refs: List[str] = []
    seen = set()
    for match in _FORMULA_REF_RE.finditer(_strip_formula_string_literals(formula)):
        sheet = match.group(1) or match.group(2) or current_sheet
        sheet = sheet.replace("''", "'").strip()
        start = _normalize_coordinate(match.group(3))
        end = _normalize_coordinate(match.group(4) or "")
        ref = f"{sheet}!{start}" if not end else f"{sheet}!{start}:{end}"
        if ref not in seen:
            seen.add(ref)
            refs.append(ref)
    return refs


def _iter_defined_names(wb: Any) -> Iterable[Any]:
    defined_names = getattr(wb, "defined_names", None)
    if defined_names is None:
        return []

    if hasattr(defined_names, "values"):
        try:
            return list(defined_names.values())
        except Exception:
            pass

    legacy = getattr(defined_names, "definedName", None)
    if legacy is not None:
        return list(legacy)
    return []


def _defined_name_destinations(wb: Any) -> List[Dict[str, str]]:
    result: List[Dict[str, str]] = []
    for defined_name in _iter_defined_names(wb):
        name = str(getattr(defined_name, "name", "") or "")
        if not name or bool(getattr(defined_name, "hidden", False)):
            continue
        try:
            destinations = list(defined_name.destinations)
        except Exception:
            destinations = []
        for sheet, ref in destinations:
            result.append({"name": name, "sheet": str(sheet), "ref": str(ref).replace("$", "")})
    return result


def _worksheet_tables(ws: Any) -> List[str]:
    result: List[str] = []
    try:
        items = list(ws.tables.items())
    except Exception:
        items = []
    for name, table in items:
        ref = getattr(table, "ref", table)
        result.append(f"{name}:{ref}")
    return result


def _add_candidate(
    candidates: Dict[Tuple[str, str], Dict[str, Any]],
    *,
    sheet: str,
    cell: str,
    reason: str,
    label: str = "",
    current_value: Any = None,
) -> None:
    if len(candidates) >= MAX_INPUT_CANDIDATES:
        return
    key = (sheet, cell)
    entry = candidates.setdefault(
        key,
        {
            "sheet": sheet,
            "cell": cell,
            "label": label,
            "current_value": _display_value(current_value),
            "reasons": [],
        },
    )
    if label and not entry.get("label"):
        entry["label"] = label
    if reason not in entry["reasons"]:
        entry["reasons"].append(reason)


def _nearby_label(ws: Any, row: int, col: int) -> str:
    for offset in (1, 2, 3):
        if col - offset >= 1:
            value = ws.cell(row=row, column=col - offset).value
            if value not in (None, "") and not _cell_is_formula(value):
                return _display_value(value)[:120]
    for offset in (1, 2):
        if row - offset >= 1:
            value = ws.cell(row=row - offset, column=col).value
            if value not in (None, "") and not _cell_is_formula(value):
                return _display_value(value)[:120]
    return ""


def _single_cell_ref(ref: str) -> bool:
    return bool(_CELL_RE.match(str(ref or "").replace("$", "")))


def _load_profile(ctx: ToolContext, profile_path: str, source: str) -> List[Dict[str, Any]]:
    if not profile_path:
        return []
    resolved, _root = _resolve_workspace_file(ctx, profile_path, source)
    data = json.loads(resolved.read_text(encoding="utf-8"))
    raw_fields = data.get("fields", data) if isinstance(data, dict) else data

    fields: List[Dict[str, Any]] = []
    if isinstance(raw_fields, dict):
        for name, cfg in raw_fields.items():
            if isinstance(cfg, str):
                fields.append({"field": name, "cell": cfg})
            elif isinstance(cfg, dict):
                fields.append({"field": name, **cfg})
    elif isinstance(raw_fields, list):
        fields = [field for field in raw_fields if isinstance(field, dict)]
    return fields


def _split_sheet_target(sheet: str, target: str) -> Tuple[str, str]:
    sheet = str(sheet or "").strip()
    target = str(target or "").strip().replace("$", "")
    if not sheet and "!" in target:
        raw_sheet, raw_target = target.rsplit("!", 1)
        sheet = raw_sheet.strip("'").replace("''", "'")
        target = raw_target
    return sheet, target


def _inspect_excel_template(
    ctx: ToolContext,
    path: str,
    source: str = "drive",
    profile_path: str = "",
) -> str:
    load_workbook, _ = _require_openpyxl()
    workbook_path, _root = _resolve_xlsx_path(ctx, path, source)
    profile_fields = _load_profile(ctx, profile_path, source) if profile_path else []

    ctx.emit_progress_fn(f"Открыл Excel-шаблон `{workbook_path.name}`. Анализирую листы, формулы и поля ввода.")
    wb = load_workbook(workbook_path, data_only=False, keep_links=True, read_only=False)
    defined_names = _defined_name_destinations(wb)
    candidates: Dict[Tuple[str, str], Dict[str, Any]] = {}
    formulas: List[Dict[str, Any]] = []

    for item in defined_names:
        if ":" in item["ref"]:
            continue
        sheet_name = item["sheet"]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cell_ref = _normalize_coordinate(item["ref"])
        cell = ws[cell_ref]
        if _cell_is_writable(ws, cell):
            _add_candidate(
                candidates,
                sheet=sheet_name,
                cell=cell_ref,
                reason=f"named_range:{item['name']}",
                label=item["name"],
                current_value=cell.value,
            )

    for field in profile_fields:
        sheet_name, cell_ref = _split_sheet_target(field.get("sheet") or "", field.get("cell") or field.get("range") or "")
        if not sheet_name or not cell_ref or ":" in cell_ref or sheet_name not in wb.sheetnames:
            continue
        if _single_cell_ref(cell_ref):
            ws = wb[sheet_name]
            cell = ws[cell_ref]
            if _cell_is_writable(ws, cell):
                _add_candidate(
                    candidates,
                    sheet=sheet_name,
                    cell=_normalize_coordinate(cell_ref),
                    reason="profile",
                    label=str(field.get("field") or field.get("name") or ""),
                    current_value=cell.value,
                )

    sheet_lines: List[str] = []
    for ws in wb.worksheets:
        max_row = min(int(ws.max_row or 1), MAX_SCAN_ROWS)
        max_col = min(int(ws.max_column or 1), MAX_SCAN_COLS)
        merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
        tables = _worksheet_tables(ws)
        sheet_lines.append(
            f"- {ws.title}: state={ws.sheet_state}, protected={'yes' if ws.protection.sheet else 'no'}, "
            f"dimension={ws.calculate_dimension()}, merged_ranges={', '.join(merged_ranges[:8]) or '-'}, "
            f"tables={', '.join(tables[:8]) or '-'}"
        )

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                value = cell.value
                if _cell_is_formula(value) and len(formulas) < MAX_INSPECT_FORMULAS:
                    deps = _formula_dependencies(str(value), ws.title)
                    formulas.append({"sheet": ws.title, "cell": cell.coordinate, "formula": value, "refs": deps[:20]})
                    for dep in deps:
                        dep_sheet, dep_ref = dep.split("!", 1)
                        if dep_sheet not in wb.sheetnames or ":" in dep_ref or not _single_cell_ref(dep_ref):
                            continue
                        dep_ws = wb[dep_sheet]
                        dep_cell = dep_ws[dep_ref]
                        if _cell_is_writable(dep_ws, dep_cell):
                            _add_candidate(
                                candidates,
                                sheet=dep_sheet,
                                cell=_normalize_coordinate(dep_ref),
                                reason=f"formula_dependency:{ws.title}!{cell.coordinate}",
                                label=_nearby_label(dep_ws, dep_cell.row, dep_cell.column),
                                current_value=dep_cell.value,
                            )
                    continue

                if value in (None, ""):
                    label = _nearby_label(ws, cell.row, cell.column)
                    if label and _cell_is_writable(ws, cell):
                        _add_candidate(
                            candidates,
                            sheet=ws.title,
                            cell=cell.coordinate,
                            reason="near_label",
                            label=label,
                            current_value=value,
                        )

    named_lines = [f"- {item['name']}: {item['sheet']}!{item['ref']}" for item in defined_names[:80]]
    formula_lines = [
        f"- {item['sheet']}!{item['cell']} {item['formula']} | refs: {', '.join(item['refs']) or '-'}"
        for item in formulas[:80]
    ]
    candidate_lines = []
    for candidate in list(candidates.values())[:MAX_INPUT_CANDIDATES]:
        confidence = "high" if any(str(r).startswith(("named_range", "profile")) for r in candidate["reasons"]) else "medium"
        candidate_lines.append(
            f"- {candidate['sheet']}!{candidate['cell']} | label={candidate['label'] or '-'} | "
            f"current={candidate['current_value'] or '<blank>'} | reasons={', '.join(candidate['reasons'])} | "
            f"suggested_confidence={confidence}"
        )

    profile_lines = []
    for field in profile_fields[:80]:
        sheet_name, target = _split_sheet_target(field.get("sheet") or "", field.get("cell") or field.get("range") or "")
        profile_lines.append(
            f"- {field.get('field') or field.get('name')}: {sheet_name}!{target} "
            f"type={field.get('type', '-')}"
        )

    return "\n".join(
        [
            "# Excel Template Inspection",
            f"- path: {path}",
            f"- workbook: {workbook_path.name}",
            f"- sheets: {len(wb.worksheets)}",
            "",
            "## Sheets",
            *(sheet_lines or ["- none"]),
            "",
            "## Named Ranges",
            *(named_lines or ["- none"]),
            "",
            "## Profile Fields",
            *(profile_lines or ["- none"]),
            "",
            "## Formula Dependencies",
            *(formula_lines or ["- none"]),
            "",
            "## Likely Input Fields",
            *(candidate_lines or ["- none"]),
            "",
            "Next: build a fill plan for the user. Ask before calling fill_excel_template for unknown or low-confidence fields.",
        ]
    )


def _coerce_updates(updates: Any) -> List[Dict[str, Any]]:
    if isinstance(updates, str):
        updates = json.loads(updates)
    if not isinstance(updates, list):
        raise ValueError("updates must be a list of objects")
    result = []
    for item in updates:
        if not isinstance(item, dict):
            raise ValueError("each update must be an object")
        result.append(item)
    return result


def _safe_cell_value(value: Any, allow_formula: bool = False) -> Any:
    if isinstance(value, str) and not allow_formula and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _set_recalc_on_open(wb: Any) -> None:
    calc = getattr(wb, "calculation", None) or getattr(wb, "calculation_properties", None)
    if calc is None:
        return
    for attr, value in (("fullCalcOnLoad", True), ("forceFullCalc", True), ("calcMode", "auto")):
        try:
            setattr(calc, attr, value)
        except Exception:
            pass


def _chart_anchor_cell(chart: Any) -> str:
    try:
        from openpyxl.utils import get_column_letter

        marker = getattr(getattr(chart, "anchor", None), "_from", None)
        if marker is not None:
            return f"{get_column_letter(int(marker.col) + 1)}{int(marker.row) + 1}"
    except Exception:
        pass
    anchor = getattr(chart, "anchor", "")
    return str(anchor or "")


def _chart_title_text(chart: Any) -> str:
    title = getattr(chart, "title", None)
    if title is None:
        return ""
    try:
        paragraphs = getattr(getattr(getattr(title, "tx", None), "rich", None), "p", []) or []
        parts: List[str] = []
        for paragraph in paragraphs:
            for run in getattr(paragraph, "r", []) or []:
                text = getattr(run, "t", "")
                if text:
                    parts.append(str(text))
        if parts:
            return "".join(parts)
    except Exception:
        pass
    return str(title)


def _chart_series_title(series: Any) -> str:
    tx = getattr(series, "tx", None)
    if tx is None:
        return ""
    try:
        if getattr(tx, "v", None):
            return str(tx.v)
        str_ref = getattr(tx, "strRef", None)
        if str_ref is not None and getattr(str_ref, "f", None):
            return str(str_ref.f)
        if getattr(tx, "rich", None):
            return str(tx.rich)
    except Exception:
        pass
    return ""


def _chart_ref_formula(obj: Any, *attrs: str) -> str:
    current = obj
    for attr in attrs:
        current = getattr(current, attr, None)
        if current is None:
            return ""
    return str(getattr(current, "f", "") or "")


def _series_value_ref(series: Any) -> str:
    return _chart_ref_formula(series, "val", "numRef")


def _series_category_ref(series: Any) -> str:
    return (
        _chart_ref_formula(series, "cat", "strRef")
        or _chart_ref_formula(series, "cat", "numRef")
        or _chart_ref_formula(series, "cat", "multiLvlStrRef")
    )


def _count_ref_values(wb: Any, default_sheet: str, ref: str, range_boundaries: Any) -> Dict[str, int]:
    try:
        sheet_name, target = _split_range_ref(default_sheet, ref)
        if sheet_name not in wb.sheetnames:
            return {"cells": 0, "nonblank": 0, "numeric": 0}
        min_col, min_row, max_col, max_row = range_boundaries(target)
    except Exception:
        return {"cells": 0, "nonblank": 0, "numeric": 0}

    ws = wb[sheet_name]
    cells = nonblank = numeric = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cells += 1
            value = cell.value
            if value in (None, ""):
                continue
            nonblank += 1
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numeric += 1
    return {"cells": cells, "nonblank": nonblank, "numeric": numeric}


def _chart_summaries(wb: Any, values_wb: Any, range_boundaries: Any) -> List[Dict[str, Any]]:
    summaries: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        for idx, chart in enumerate(getattr(ws, "_charts", []) or [], start=1):
            anchor = _chart_anchor_cell(chart)
            series_summaries: List[Dict[str, Any]] = []
            for series in list(getattr(chart, "series", []) or [])[:MAX_CHART_SERIES]:
                value_ref = _series_value_ref(series)
                category_ref = _series_category_ref(series)
                counts = _count_ref_values(values_wb, ws.title, value_ref, range_boundaries)
                series_summaries.append(
                    {
                        "title": _chart_series_title(series),
                        "values": value_ref,
                        "categories": category_ref,
                        **counts,
                    }
                )

            warnings: List[str] = []
            if ws.sheet_state != "visible":
                warnings.append("sheet_hidden")
            if not series_summaries:
                warnings.append("no_series")
            if any(item.get("numeric", 0) == 0 for item in series_summaries):
                warnings.append("series_without_numeric_values")
            try:
                from openpyxl.utils.cell import coordinate_to_tuple

                row, col = coordinate_to_tuple(anchor)
                if col > int(ws.max_column or 1) + 1 or row > int(ws.max_row or 1) + 5:
                    warnings.append("anchor_outside_used_cell_range")
            except Exception:
                pass

            summaries.append(
                {
                    "sheet": ws.title,
                    "index": idx,
                    "type": type(chart).__name__,
                    "title": _chart_title_text(chart),
                    "anchor": anchor,
                    "series_count": len(getattr(chart, "series", []) or []),
                    "series": series_summaries,
                    "warnings": warnings,
                }
            )
    return summaries


def _format_chart_summaries(summaries: List[Dict[str, Any]]) -> List[str]:
    if not summaries:
        return ["- charts: 0", "- status: no native Excel charts found"]

    lines = [f"- charts: {len(summaries)}"]
    for item in summaries:
        warnings = ",".join(item["warnings"]) if item["warnings"] else "-"
        title = item["title"] or "-"
        lines.append(
            f"- {item['sheet']}!{item['anchor']} | type={item['type']} | title={title} | "
            f"series_count={item['series_count']} | warnings={warnings}"
        )
        for series in item["series"]:
            series_title = series["title"] or "-"
            lines.append(
                f"  - series={series_title} | values={series['values'] or '-'} | "
                f"categories={series['categories'] or '-'} | numeric_points={series['numeric']}/{series['cells']}"
            )
    return lines


def _merged_anchor(ws: Any, coordinate: str) -> str:
    for merged_range in ws.merged_cells.ranges:
        if coordinate in merged_range:
            return merged_range.start_cell.coordinate
    return coordinate


def _target_cells(ws: Any, target: str, range_boundaries: Any) -> List[Any]:
    target = str(target or "").replace("$", "")
    if ":" not in target:
        return [ws[_normalize_coordinate(target)]]

    min_col, min_row, max_col, max_row = range_boundaries(target)
    cells = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        cells.extend(row)
    return cells


def _value_for_target(value: Any, index: int, cell_count: int) -> Any:
    if cell_count == 1:
        return value
    if not isinstance(value, list):
        return value

    flat: List[Any] = []
    for row in value:
        if isinstance(row, list):
            flat.extend(row)
        else:
            flat.append(row)
    if index >= len(flat):
        raise ValueError("range update value list is smaller than target range")
    return flat[index]


def _fill_excel_template(
    ctx: ToolContext,
    path: str,
    updates: Any,
    output_path: str = "",
    overwrite: bool = False,
    send_to_chat: bool = True,
) -> str:
    load_workbook, range_boundaries = _require_openpyxl()
    workbook_path, root = _resolve_xlsx_path(ctx, path, "drive")
    output, rel_output = _safe_output_path(root, output_path, workbook_path.name, bool(overwrite))
    update_items = _coerce_updates(updates)

    ctx.emit_progress_fn(f"Заполняю копию Excel-шаблона `{workbook_path.name}`: {len(update_items)} обновлений.")
    wb = load_workbook(workbook_path, data_only=False, keep_links=True, read_only=False)
    written: List[str] = []
    rejected: List[str] = []

    for idx, update in enumerate(update_items, start=1):
        sheet_name = str(update.get("sheet") or "").strip()
        target = str(update.get("cell") or update.get("range") or "").strip()
        field = str(update.get("field") or update.get("label") or f"update_{idx}")
        if not sheet_name or not target:
            rejected.append(f"{field}: missing sheet/cell")
            continue
        if sheet_name not in wb.sheetnames:
            rejected.append(f"{field}: sheet not found: {sheet_name}")
            continue

        confidence = update.get("confidence")
        try:
            confidence_value = float(confidence) if confidence is not None else None
        except Exception:
            confidence_value = None
        if confidence_value is not None and confidence_value < MIN_WRITE_CONFIDENCE and not bool(update.get("confirmed")):
            rejected.append(f"{field}: low confidence {confidence_value:.2f}; ask user or pass confirmed=true")
            continue

        ws = wb[sheet_name]
        try:
            cells = _target_cells(ws, target, range_boundaries)
        except Exception as exc:
            rejected.append(f"{field}: invalid target {target}: {exc}")
            continue

        for pos, cell in enumerate(cells):
            anchor = _merged_anchor(ws, cell.coordinate)
            if anchor != cell.coordinate:
                rejected.append(f"{field}: {sheet_name}!{cell.coordinate} is inside merged range anchored at {anchor}")
                continue
            if _cell_is_formula(cell.value):
                rejected.append(f"{field}: refusing to overwrite formula at {sheet_name}!{cell.coordinate}")
                continue
            if ws.protection.sheet and getattr(getattr(cell, "protection", None), "locked", False):
                rejected.append(f"{field}: refusing to overwrite locked protected cell {sheet_name}!{cell.coordinate}")
                continue

            try:
                cell.value = _safe_cell_value(
                    _value_for_target(update.get("value"), pos, len(cells)),
                    allow_formula=bool(update.get("allow_formula")),
                )
            except Exception as exc:
                rejected.append(f"{field}: could not write {sheet_name}!{cell.coordinate}: {exc}")
                continue
            written.append(f"{sheet_name}!{cell.coordinate} ({field})")

    if not written:
        lines = [
            "⚠️ No Excel updates were written",
            f"- source_path: {path}",
            f"- rejected: {len(rejected)}",
            "- output_path: not_created",
            "- telegram_delivery: skipped",
        ]
        if rejected:
            lines.append("## Rejected")
            lines.extend(f"- {item}" for item in rejected[:80])
        return "\n".join(lines)

    queued = False
    _set_recalc_on_open(wb)
    wb.save(output)
    if send_to_chat and ctx.current_chat_id:
        ctx.pending_events.append(
            {
                "type": "send_document",
                "chat_id": ctx.current_chat_id,
                "path": rel_output,
                "caption": "Заполненный Excel-шаблон",
                "filename": output.name,
                "mime_type": XLSX_MIME_TYPE,
                **_scope(ctx),
            }
        )
        queued = True

    lines = [
        "OK: Excel template filled",
        f"- source_path: {path}",
        f"- output_path: {rel_output}",
        f"- written: {len(written)}",
        f"- rejected: {len(rejected)}",
        f"- telegram_delivery: {'queued' if queued else 'skipped'}",
    ]
    if written:
        lines.append("## Written")
        lines.extend(f"- {item}" for item in written[:80])
    if rejected:
        lines.append("## Rejected")
        lines.extend(f"- {item}" for item in rejected[:80])
    return "\n".join(lines)


def _inspect_excel_charts(
    ctx: ToolContext,
    path: str,
    source: str = "drive",
) -> str:
    load_workbook, range_boundaries = _require_openpyxl()
    workbook_path, _root = _resolve_xlsx_path(ctx, path, source)

    ctx.emit_progress_fn(f"Проверяю графики в Excel-файле `{workbook_path.name}`.")
    wb = load_workbook(workbook_path, data_only=False, keep_links=True, read_only=False)
    values_wb = load_workbook(workbook_path, data_only=True, keep_links=True, read_only=False)
    summaries = _chart_summaries(wb, values_wb, range_boundaries)

    lines = [
        "# Excel Chart Inspection",
        f"- path: {path}",
        f"- workbook: {workbook_path.name}",
        f"- sheets: {len(wb.worksheets)}",
        "",
        "## Charts",
        *_format_chart_summaries(summaries),
    ]
    if any(item["warnings"] for item in summaries):
        lines.extend(
            [
                "",
                "## Guidance",
                "- For user-facing files, place the main chart on a visible sheet near A1 or on a dedicated chart sheet.",
                "- Do not tell the user a chart was created until this inspection reports a chart with series and numeric points.",
            ]
        )
    return "\n".join(lines)


def _create_excel_line_chart(
    ctx: ToolContext,
    path: str,
    data_sheet: str,
    category_range: str,
    value_ranges: Any,
    series_names: Any = None,
    title: str = "Line chart",
    output_path: str = "",
    chart_sheet: str = "Chart",
    anchor: str = "A3",
    y_axis_title: str = "",
    x_axis_title: str = "",
    percent_axis: bool = False,
    overwrite: bool = False,
    send_to_chat: bool = True,
) -> str:
    load_workbook, range_boundaries = _require_openpyxl()
    from openpyxl.chart import LineChart, Reference, Series

    workbook_path, root = _resolve_xlsx_path(ctx, path, "drive")
    output, rel_output = _safe_output_path(root, output_path, workbook_path.name, bool(overwrite), "chart")

    value_ranges_list = value_ranges
    if isinstance(value_ranges_list, str):
        value_ranges_list = [item.strip() for item in value_ranges_list.split(",") if item.strip()]
    if not isinstance(value_ranges_list, list) or not value_ranges_list:
        raise ValueError("value_ranges must be a non-empty list of ranges")
    if len(value_ranges_list) > MAX_CHART_SERIES:
        raise ValueError(f"value_ranges supports at most {MAX_CHART_SERIES} series")

    names_list = series_names
    if isinstance(names_list, str):
        names_list = [item.strip() for item in names_list.split(",")]
    if names_list is None:
        names_list = []
    if not isinstance(names_list, list):
        raise ValueError("series_names must be a list when supplied")

    ctx.emit_progress_fn(f"Создаю видимый линейный график в Excel-файле `{workbook_path.name}`.")
    wb = load_workbook(workbook_path, data_only=False, keep_links=True, read_only=False)
    if data_sheet not in wb.sheetnames:
        raise ValueError(f"data_sheet not found: {data_sheet}")
    ws = wb[data_sheet]

    category_sheet, category_target = _split_range_ref(data_sheet, category_range)
    if category_sheet != data_sheet:
        raise ValueError("category_range must point to data_sheet in this tool version")
    category_rows, category_cols, category_cells = _range_dimensions(category_target, range_boundaries)
    if category_cols != 1:
        raise ValueError("category_range must be a single column")
    if category_cells > MAX_CHART_POINTS:
        raise ValueError(f"category_range is too large; max {MAX_CHART_POINTS} points")

    category_min_col, category_min_row, category_max_col, category_max_row = range_boundaries(category_target)
    categories = Reference(
        ws,
        min_col=category_min_col,
        min_row=category_min_row,
        max_col=category_max_col,
        max_row=category_max_row,
    )

    chart = LineChart()
    chart.title = str(title or "Line chart")[:255]
    chart.style = 13
    chart.height = 12
    chart.width = 24
    chart.legend.position = "r"
    chart.y_axis.title = str(y_axis_title or "")
    chart.x_axis.title = str(x_axis_title or "")
    if percent_axis:
        chart.y_axis.numFmt = "0%"

    for idx, raw_ref in enumerate(value_ranges_list):
        value_sheet, value_target = _split_range_ref(data_sheet, str(raw_ref))
        if value_sheet != data_sheet:
            raise ValueError("value_ranges must point to data_sheet in this tool version")
        value_rows, value_cols, value_cells = _range_dimensions(value_target, range_boundaries)
        if value_cols != 1:
            raise ValueError(f"value range must be a single column: {raw_ref}")
        if value_rows != category_rows:
            raise ValueError(
                f"value range {raw_ref} has {value_rows} rows but category_range has {category_rows}"
            )
        if value_cells > MAX_CHART_POINTS:
            raise ValueError(f"value range is too large; max {MAX_CHART_POINTS} points: {raw_ref}")
        min_col, min_row, max_col, max_row = range_boundaries(value_target)
        values = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
        series_title = str(names_list[idx]) if idx < len(names_list) and names_list[idx] not in (None, "") else ""
        if not series_title and min_row > 1:
            header = ws.cell(row=min_row - 1, column=min_col).value
            series_title = _display_value(header)
        chart.series.append(Series(values, title=series_title or f"Series {idx + 1}"))

    chart.set_categories(categories)

    target_sheet_name = _safe_sheet_title(chart_sheet or data_sheet)
    if target_sheet_name in wb.sheetnames:
        chart_ws = wb[target_sheet_name]
        chart_ws._charts = []
    else:
        chart_ws = wb.create_sheet(_unique_sheet_title(wb, target_sheet_name), 0)
    chart_ws.sheet_state = "visible"
    chart_ws["A1"] = str(title or "Line chart")
    chart_ws["A1"].style = "Title"
    chart_ws.add_chart(chart, str(anchor or "A3"))
    chart_ws.freeze_panes = "A3"

    _set_recalc_on_open(wb)
    wb.save(output)

    queued = False
    if send_to_chat and ctx.current_chat_id:
        ctx.pending_events.append(
            {
                "type": "send_document",
                "chat_id": ctx.current_chat_id,
                "path": rel_output,
                "caption": "Excel-файл с графиком",
                "filename": output.name,
                "mime_type": XLSX_MIME_TYPE,
                **_scope(ctx),
            }
        )
        queued = True

    verify_wb = load_workbook(output, data_only=False, keep_links=True, read_only=False)
    values_wb = load_workbook(output, data_only=True, keep_links=True, read_only=False)
    summaries = _chart_summaries(verify_wb, values_wb, range_boundaries)
    lines = [
        "OK: Excel line chart created",
        f"- source_path: {path}",
        f"- output_path: {rel_output}",
        f"- chart_sheet: {chart_ws.title}",
        f"- data_sheet: {data_sheet}",
        f"- category_range: {_quote_sheet_for_formula(data_sheet)}!{category_target}",
        f"- value_ranges: {len(value_ranges_list)}",
        f"- telegram_delivery: {'queued' if queued else 'skipped'}",
        "",
        "## Chart Verification",
        *_format_chart_summaries(summaries),
    ]
    return "\n".join(lines)


def get_tools() -> List[ToolEntry]:
    return [
        ToolEntry(
            "inspect_excel_template",
            {
                "name": "inspect_excel_template",
                "description": (
                    "Inspect a .xlsx workbook as a fillable Excel template. Use this when the user wants "
                    "a spreadsheet template filled, not merely summarized. It reports sheets, formulas, "
                    "formula dependencies, named ranges, tables, merged ranges, and likely input cells. "
                    "After inspection, show the user a fill plan and ask about uncertain fields before writing."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the .xlsx template in Drive or repo."},
                        "source": {
                            "type": "string",
                            "enum": ["drive", "repo"],
                            "default": "drive",
                            "description": "Where to read the template from. repo is admin-only.",
                        },
                        "profile_path": {
                            "type": "string",
                            "description": "Optional JSON profile mapping business fields to sheet/cell/type.",
                        },
                    },
                    "required": ["path"],
                },
            },
            _inspect_excel_template,
            timeout_sec=60,
        ),
        ToolEntry(
            "fill_excel_template",
            {
                "name": "fill_excel_template",
                "description": (
                    "Fill confirmed values into a copy of a .xlsx template and optionally send it to Telegram. "
                    "Use only after inspect_excel_template and a user-approved fill plan. The tool preserves "
                    "formatting and formulas, refuses to overwrite formula cells, and rejects low-confidence "
                    "updates unless confirmed=true is supplied for that update."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Drive path to the source .xlsx template."},
                        "updates": {
                            "type": "array",
                            "description": (
                                "Confirmed writes. Each item supports sheet, cell or range, value, field, "
                                "confidence, confirmed, and allow_formula."
                            ),
                            "items": {
                                "type": "object",
                                "properties": {
                                    "sheet": {"type": "string"},
                                    "cell": {"type": "string"},
                                    "range": {"type": "string"},
                                    "value": {},
                                    "field": {"type": "string"},
                                    "confidence": {"type": "number"},
                                    "confirmed": {"type": "boolean"},
                                    "allow_formula": {"type": "boolean"},
                                },
                            },
                        },
                        "output_path": {
                            "type": "string",
                            "description": "Optional output .xlsx path in Drive. Defaults to spreadsheets/<name>-filled-<timestamp>.xlsx.",
                        },
                        "overwrite": {
                            "type": "boolean",
                            "default": False,
                            "description": "Overwrite output_path if it exists; otherwise create a deduplicated filename.",
                        },
                        "send_to_chat": {
                            "type": "boolean",
                            "default": True,
                            "description": "Queue the resulting .xlsx for Telegram delivery when there is an active chat.",
                        },
                    },
                    "required": ["path", "updates"],
                },
            },
            _fill_excel_template,
            timeout_sec=60,
        ),
        ToolEntry(
            "inspect_excel_charts",
            {
                "name": "inspect_excel_charts",
                "description": (
                    "Inspect native charts embedded in a .xlsx workbook. Use this before telling a user "
                    "that an Excel chart/diagram was created or when a user says a chart is missing. "
                    "It reports chart sheet, anchor cell, series count, source ranges, numeric point counts, "
                    "and visibility warnings such as hidden sheets or anchors outside the used cell range."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the .xlsx workbook in Drive or repo."},
                        "source": {
                            "type": "string",
                            "enum": ["drive", "repo"],
                            "default": "drive",
                            "description": "Where to read the workbook from. repo is admin-only.",
                        },
                    },
                    "required": ["path"],
                },
            },
            _inspect_excel_charts,
            timeout_sec=60,
        ),
        ToolEntry(
            "create_excel_line_chart",
            {
                "name": "create_excel_line_chart",
                "description": (
                    "Create a visible native Excel line chart from an existing table in a .xlsx workbook. "
                    "Use after preparing chart-ready data, especially when the user explicitly asks to "
                    "build a graph/diagram/chart in Excel. By default it creates or refreshes a visible "
                    "'Chart' worksheet and anchors the chart near A1, then verifies the resulting workbook."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the source .xlsx workbook in Drive."},
                        "data_sheet": {"type": "string", "description": "Worksheet containing the chart-ready data."},
                        "category_range": {
                            "type": "string",
                            "description": "Single-column category labels, e.g. A2:A53 or LFL!A2:A53.",
                        },
                        "value_ranges": {
                            "type": "array",
                            "description": "One or more single-column numeric ranges with the same row count as category_range.",
                            "items": {"type": "string"},
                        },
                        "series_names": {
                            "type": "array",
                            "description": "Optional series labels matching value_ranges.",
                            "items": {"type": "string"},
                        },
                        "title": {"type": "string", "default": "Line chart"},
                        "output_path": {
                            "type": "string",
                            "description": "Optional output .xlsx path in Drive. Defaults to spreadsheets/<name>-chart-<timestamp>.xlsx.",
                        },
                        "chart_sheet": {
                            "type": "string",
                            "default": "Chart",
                            "description": "Visible worksheet where the chart will be placed. Existing charts on this sheet are refreshed.",
                        },
                        "anchor": {"type": "string", "default": "A3", "description": "Chart anchor cell on chart_sheet."},
                        "x_axis_title": {"type": "string"},
                        "y_axis_title": {"type": "string"},
                        "percent_axis": {
                            "type": "boolean",
                            "default": False,
                            "description": "Format the y-axis as percentages.",
                        },
                        "overwrite": {
                            "type": "boolean",
                            "default": False,
                            "description": "Overwrite output_path if it exists; otherwise create a deduplicated filename.",
                        },
                        "send_to_chat": {
                            "type": "boolean",
                            "default": True,
                            "description": "Queue the resulting .xlsx for Telegram delivery when there is an active chat.",
                        },
                    },
                    "required": ["path", "data_sheet", "category_range", "value_ranges"],
                },
            },
            _create_excel_line_chart,
            timeout_sec=60,
        ),
    ]
