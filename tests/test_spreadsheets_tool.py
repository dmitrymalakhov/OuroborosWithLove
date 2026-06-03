import json
import pathlib

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

from ouroboros.tools.registry import ToolContext, ToolRegistry
from ouroboros.tools.spreadsheets import (
    XLSX_MIME_TYPE,
    _fill_excel_template,
    _formula_dependencies,
    _inspect_excel_template,
)


def _ctx(repo_dir: pathlib.Path, drive_root: pathlib.Path) -> ToolContext:
    repo_dir.mkdir(parents=True, exist_ok=True)
    drive_root.mkdir(parents=True, exist_ok=True)
    return ToolContext(
        repo_dir=repo_dir,
        drive_root=drive_root,
        current_chat_id=123,
        current_user_id=456,
        user_role="user",
    )


def _add_defined_name(wb: Workbook, name: str, attr_text: str) -> None:
    defined_name = DefinedName(name, attr_text=attr_text)
    if hasattr(wb.defined_names, "add"):
        wb.defined_names.add(defined_name)
    else:
        wb.defined_names.append(defined_name)


def _write_template(path: pathlib.Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Company"
    ws["A2"] = "Revenue"
    ws["B1"] = None
    ws["B2"] = None
    ws["C2"] = "=B2*0.2"
    ws["D1"] = "Merged label"
    ws.merge_cells("D1:E1")
    ws["A5"] = "Metric"
    ws["B5"] = "Value"
    ws["A6"] = "Headcount"
    ws["B6"] = None
    table = Table(displayName="InputTable", ref="A5:B6")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showRowStripes=True)
    ws.add_table(table)
    ws["B1"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws.column_dimensions["B"].width = 24

    calc = wb.create_sheet("Calc")
    calc["A1"] = "=Inputs!B2+10"
    calc.sheet_state = "hidden"
    _add_defined_name(wb, "company_name", "'Inputs'!$B$1")
    wb.save(path)


def test_inspect_excel_template_reports_structure_and_input_candidates(tmp_path):
    drive = tmp_path / "drive"
    repo = tmp_path / "repo"
    drive.mkdir()
    _write_template(drive / "template.xlsx")
    (drive / "profile.json").write_text(
        json.dumps({"fields": {"company": "Inputs!B1", "revenue": {"sheet": "Inputs", "cell": "B2", "type": "number"}}}),
        encoding="utf-8",
    )

    result = _inspect_excel_template(_ctx(repo, drive), path="template.xlsx", profile_path="profile.json")

    assert "# Excel Template Inspection" in result
    assert "Inputs: state=visible" in result
    assert "Calc: state=hidden" in result
    assert "merged_ranges=D1:E1" in result
    assert "InputTable" in result
    assert "company_name: Inputs!B1" in result
    assert "company: Inputs!B1" in result
    assert "revenue: Inputs!B2 type=number" in result
    assert "Inputs!C2 =B2*0.2" in result
    assert "Calc!A1 =Inputs!B2+10" in result
    assert "Inputs!B1" in result
    assert "Inputs!B2" in result


def test_fill_excel_template_writes_copy_preserves_template_and_queues_delivery(tmp_path):
    drive = tmp_path / "drive"
    repo = tmp_path / "repo"
    drive.mkdir()
    _write_template(drive / "template.xlsx")
    ctx = _ctx(repo, drive)

    result = _fill_excel_template(
        ctx,
        path="template.xlsx",
        output_path="exports/filled.xlsx",
        updates=[
            {"sheet": "Inputs", "cell": "B1", "field": "Company", "value": "ACME", "confidence": 0.95},
            {"sheet": "Inputs", "cell": "B2", "field": "Revenue", "value": 1000, "confidence": 0.95},
            {"sheet": "Inputs", "cell": "C2", "field": "Formula cell", "value": 1, "confidence": 0.99},
            {"sheet": "Inputs", "cell": "B6", "field": "Headcount", "value": 12, "confidence": 0.4},
            {"sheet": "Inputs", "cell": "A3", "field": "Formula injection", "value": "=1+1", "confidence": 0.95},
        ],
    )

    assert "OK: Excel template filled" in result
    assert "- written: 3" in result
    assert "- rejected: 2" in result
    assert "refusing to overwrite formula" in result
    assert "low confidence" in result

    original = load_workbook(drive / "template.xlsx", data_only=False)
    filled = load_workbook(drive / "exports" / "filled.xlsx", data_only=False)
    assert original["Inputs"]["B1"].value is None
    assert filled["Inputs"]["B1"].value == "ACME"
    assert filled["Inputs"]["B2"].value == 1000
    assert filled["Inputs"]["C2"].value == "=B2*0.2"
    assert filled["Inputs"]["A3"].value == "'=1+1"
    assert filled["Inputs"]["B1"].fill.fgColor.rgb == "00FFF2CC"
    assert filled["Inputs"].column_dimensions["B"].width == 24
    assert "D1:E1" in [str(rng) for rng in filled["Inputs"].merged_cells.ranges]

    assert len(ctx.pending_events) == 1
    event = ctx.pending_events[0]
    assert event["type"] == "send_document"
    assert event["path"] == "exports/filled.xlsx"
    assert event["mime_type"] == XLSX_MIME_TYPE


def test_fill_excel_template_does_not_create_output_when_nothing_written(tmp_path):
    drive = tmp_path / "drive"
    repo = tmp_path / "repo"
    drive.mkdir()
    _write_template(drive / "template.xlsx")
    ctx = _ctx(repo, drive)

    result = _fill_excel_template(
        ctx,
        path="template.xlsx",
        output_path="exports/unchanged.xlsx",
        updates=[
            {"sheet": "Inputs", "cell": "C2", "field": "Formula cell", "value": 1, "confidence": 0.99},
            {"sheet": "Inputs", "cell": "B6", "field": "Headcount", "value": 12, "confidence": 0.4},
        ],
    )

    assert "No Excel updates were written" in result
    assert not (drive / "exports" / "unchanged.xlsx").exists()
    assert ctx.pending_events == []


def test_formula_dependencies_ignore_cell_like_text_inside_strings():
    refs = _formula_dependencies('=IF(A1>0,"Use B2 text",Sheet2!C3)', "Inputs")

    assert refs == ["Inputs!A1", "Sheet2!C3"]


def test_excel_template_tools_reject_path_traversal(tmp_path):
    drive = tmp_path / "drive"
    repo = tmp_path / "repo"
    drive.mkdir()
    _write_template(drive / "template.xlsx")

    with pytest.raises(ValueError, match="Path traversal"):
        _inspect_excel_template(_ctx(repo, drive), path="../template.xlsx")


def test_protected_locked_cells_are_not_suggested_or_written(tmp_path):
    drive = tmp_path / "drive"
    repo = tmp_path / "repo"
    drive.mkdir()
    path = drive / "protected.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Protected"
    ws["A1"] = "Locked input"
    ws["B1"] = None
    ws.protection.sheet = True
    wb.save(path)
    ctx = _ctx(repo, drive)

    inspect_result = _inspect_excel_template(ctx, path="protected.xlsx")
    fill_result = _fill_excel_template(
        ctx,
        path="protected.xlsx",
        output_path="exports/protected-filled.xlsx",
        updates=[{"sheet": "Protected", "cell": "B1", "field": "Locked input", "value": "x", "confidence": 0.95}],
    )

    assert "Protected!B1" not in inspect_result
    assert "locked protected cell" in fill_result
    assert not (drive / "exports" / "protected-filled.xlsx").exists()
    assert ctx.pending_events == []


def test_spreadsheets_pack_is_registered(tmp_path):
    registry = ToolRegistry(repo_dir=tmp_path / "repo", drive_root=tmp_path / "drive")

    tools = registry.get_tools_by_pack("excel", include_dependencies=True)

    assert "inspect_excel_template" in tools
    assert "fill_excel_template" in tools
    assert "analyze_document" in tools
